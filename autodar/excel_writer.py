from __future__ import annotations

import logging
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from .errors import ReportError
from .models import ReportData, TemplateConfig

LOGGER = logging.getLogger(__name__)


def _set_date(cell, value) -> None:
    number_format = cell.number_format
    cell.value = value
    if number_format and number_format != "General":
        cell.number_format = number_format
    else:
        cell.number_format = "dd/mm/yyyy"


def _assert_writable_cells(sheet, addresses: list[str]) -> None:
    for address in addresses:
        if isinstance(sheet[address], MergedCell):
            raise ReportError(f"Configured cell {address} is not the top-left cell of its merged range.")


def write_report(template: TemplateConfig, data: ReportData, destination: Path) -> None:
    if not template.file.is_file():
        raise ReportError(f"Excel template was not found: {template.file}")
    if destination.resolve() == template.file.resolve():
        raise ReportError("Choose a different file name. The original template cannot be overwritten.")
    try:
        workbook = load_workbook(template.file, data_only=False)
    except Exception as exc:
        LOGGER.exception("Could not open workbook %s", template.file)
        raise ReportError(f"The Excel template could not be opened: {template.file.name}") from exc
    try:
        if template.sheet not in workbook.sheetnames:
            raise ReportError(f"Worksheet '{template.sheet}' was not found in the template.")
        sheet = workbook[template.sheet]
        cells = template.cells
        addresses = [
            cells.shift_date,
            cells.team_leader.name,
            cells.team_leader.pin,
            *(address for person in cells.security_officers for address in (person.name, person.pin)),
            cells.control_room.name,
            cells.control_room.pin,
            *cells.patrols,
            cells.dar_officer,
            cells.dar_date,
        ]
        _assert_writable_cells(sheet, addresses)
        _set_date(sheet[cells.shift_date], data.shift_date)
        sheet[cells.team_leader.name] = data.team_leader.name
        sheet[cells.team_leader.pin] = data.team_leader.pin
        for mapping, employee in zip(cells.security_officers, data.security_officers, strict=True):
            sheet[mapping.name] = employee.name
            sheet[mapping.pin] = employee.pin
        sheet[cells.control_room.name] = data.control_room.name
        sheet[cells.control_room.pin] = data.control_room.pin
        for address, employee in zip(cells.patrols, data.patrols, strict=True):
            sheet[address] = f"{template.patrol_prefix} {employee.name}"
        sheet[cells.dar_officer] = data.dar_officer.name
        _set_date(sheet[cells.dar_date], data.dar_date)
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
    except ReportError:
        raise
    except PermissionError as exc:
        LOGGER.exception("Permission error saving report to %s", destination)
        raise ReportError("The report could not be saved. Close it in Excel or choose another location.") from exc
    except OSError as exc:
        LOGGER.exception("OS error saving report to %s", destination)
        raise ReportError(f"The report could not be saved to: {destination}") from exc
    except Exception as exc:
        LOGGER.exception("Unexpected Excel error")
        raise ReportError("An unexpected Excel error occurred while generating the report.") from exc
    finally:
        workbook.close()

