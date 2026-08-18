from datetime import date
from hashlib import sha256
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from autodar.config import load_templates
from autodar.employees import load_employees
from autodar.ooxml_writer import write_report
from autodar.models import ReportData


ROOT = Path(__file__).resolve().parents[1]


@pytest.mark.parametrize(
    ("template_key", "control_row", "patrol_rows", "dar_row"),
    [
        ("morning", 27, (30, 31, 32), 35),
        ("evening", 28, (31, 32, 33), 36),
        ("night", 27, (30, 31, 32), 35),
        ("weekend", 27, (30, 31, 32), 35),
    ],
)
def test_shift_report_mapping_and_template_integrity(tmp_path, template_key, control_row, patrol_rows, dar_row):
    employees = list(load_employees(ROOT / "names_by_number.json").values())
    template = load_templates(ROOT / "templates.json", ROOT)[template_key]
    original_hash = sha256(template.file.read_bytes()).hexdigest()
    data = ReportData(
        shift_date=date(2026, 8, 19),
        team_leader=employees[0],
        security_officers=tuple(employees[1:4]),
        control_room=employees[4],
        patrol_officer=employees[1],
        dar_officer=employees[0],
    )
    output = tmp_path / f"{template_key}-report.xlsx"
    write_report(template, data, output)

    original = load_workbook(template.file, data_only=False)
    generated = load_workbook(output, data_only=False)
    sheet = generated["Shift Report"]
    assert sheet["H9"].value == date(2026, 8, 19)
    assert sheet["H9"].number_format == "dd/mm/yyyy"
    assert (sheet["D17"].value, sheet["E17"].value) == (employees[0].name, employees[0].pin)
    for row, employee in zip((23, 24, 25), employees[1:4], strict=True):
        assert (sheet[f"D{row}"].value, sheet[f"E{row}"].value) == (employee.name, employee.pin)
    assert (sheet[f"D{control_row}"].value, sheet[f"E{control_row}"].value) == (employees[4].name, employees[4].pin)
    for row in patrol_rows:
        assert sheet[f"D{row}"].value == f"Patrol by security officer: {employees[1].name}"
    assert sheet[f"D{dar_row}"].value == employees[0].name
    assert sheet[f"M{dar_row}"].value == date(2026, 8, 19)
    assert sheet[f"M{dar_row}"].number_format == "dd/mm/yyyy"
    if template_key == "evening":
        assert (sheet["D27"].value, sheet["E27"].value) == ("Name", "Pin")
    assert generated.sheetnames == original.sheetnames
    assert set(generated["Shift Report"].merged_cells.ranges) == set(original["Shift Report"].merged_cells.ranges)
    for coordinate in ("F17", "H17", "J17", "L17", "F23", "H23", "J23", "L23"):
        assert sheet[coordinate].value == original["Shift Report"][coordinate].value
        assert sheet[coordinate]._style == original["Shift Report"][coordinate]._style
    original.close()
    generated.close()
    with ZipFile(template.file) as source, ZipFile(output) as result:
        assert set(source.namelist()) == set(result.namelist())
        for name in source.namelist():
            if name not in {"xl/worksheets/sheet1.xml", "xl/styles.xml"}:
                assert source.read(name) == result.read(name)
        worksheet_xml = result.read("xl/worksheets/sheet1.xml").decode("utf-8")
        for prefix in ("mc", "x14ac", "xr", "xr2", "xr3"):
            assert f"xmlns:{prefix}=" in worksheet_xml
        assert 'mc:Ignorable="x14ac xr xr2 xr3"' in worksheet_xml
    assert sha256(template.file.read_bytes()).hexdigest() == original_hash
